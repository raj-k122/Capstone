import cv2
import mediapipe as mp
import openpyxl  # For Excel manipulation

mp_pose = mp.solutions.pose
pose = mp_pose.Pose()

video_name = "P7_E3T1_Front"
cap = cv2.VideoCapture(f"videos/original footage/{video_name}.mp4")

landmark_data = {}
landmark_ids = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32] # Example: collect data for all landmarks. You can change this.

for id in landmark_ids:
    landmark_data[id] = {'x': [], 'y': []}

# Create a new Excel workbook
workbook = openpyxl.Workbook()

while cap.isOpened():
    ret, frame = cap.read()
    if not ret:
        break

    frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    results = pose.process(frame_rgb)

    if results.pose_landmarks:
        for i, landmark in enumerate(results.pose_landmarks.landmark):
            if i in landmark_ids:
                h, w, c = frame.shape
                x = int(landmark.x * w)
                y = int(landmark.y * h)
                cv2.circle(frame, (x, y), 5, (0, 0, 255), -1)

                landmark_data[i]['x'].append(x)
                landmark_data[i]['y'].append(y)

    cv2.imshow("Video with Pose Estimation", frame)

    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

cap.release()
cv2.destroyAllWindows()


# Write data to Excel sheets
# for landmark_id, data in landmark_data.items():
#     sheet_name = f"Landmark_{landmark_id}"  # Create sheet name
#     if sheet_name in workbook.sheetnames: # Check if the sheet exists
#         sheet = workbook[sheet_name] # If exists, get the sheet
#     else:
#         sheet = workbook.create_sheet(title=sheet_name)  # Create a new sheet
    
#     # Add header row (only once per sheet)
#     if sheet.max_row is None: # Check if the sheet is empty
#         sheet.append(["Frame", "X", "Y"])  # Add header

#     # Append data rows
#     for j in range(len(data['x'])):  # Iterate through the data points
#         sheet.append([j+1, data['x'][j], data['y'][j]]) # j+1 because frames start at 1

# # Save the workbook
# workbook.save(f"generated/patients/p3/{video_name}.xlsx")

# print("Data written to xlsx")